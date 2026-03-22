"""Excel model builder — generates the canonical 7-sheet financial model.

Pure utility. Uses openpyxl only. No database calls.
See instructions/domain/models.md for the canonical template spec.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Formatting constants
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_BASE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_BULL_FILL = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
_BEAR_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
_BOLD = Font(bold=True)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)

_FMT_DOLLAR = '#,##0.0'
_FMT_PCT = '0.0%'
_FMT_PER_SHARE = '#,##0.00'
_FMT_INT = '#,##0'
_FMT_RATIO = '0.00'

# ---------------------------------------------------------------------------
# Row definitions per sheet (exact names from models.md)
# ---------------------------------------------------------------------------
ASSUMPTIONS_ROWS = [
    "Revenue Growth Rate (%)",
    "Gross Margin (%)",
    "R&D as % of Revenue",
    "S&M as % of Revenue",
    "G&A as % of Revenue",
    "EBITDA Margin (%)",
    "Effective Tax Rate (%)",
    "D&A as % of Revenue",
    "Capex as % of Revenue",
    "Days Sales Outstanding",
    "Days Inventory Outstanding",
    "Days Payable Outstanding",
    "Share Count - Basic (M)",
    "Share Count - Diluted (M)",
    "WACC (%)",
    "Terminal Growth Rate (%)",
]

INCOME_STATEMENT_ROWS = [
    "Revenue", "Cost of Revenue", "Gross Profit", "Gross Margin %",
    "R&D", "S&M", "G&A", "Total Operating Expenses",
    "EBIT", "EBIT Margin %", "D&A", "EBITDA", "EBITDA Margin %",
    "Interest Expense", "Other Income / (Expense)", "Pre-tax Income",
    "Income Tax", "Net Income", "Net Margin %",
    "Shares Outstanding (basic)", "Shares Outstanding (diluted)",
    "EPS (basic)", "EPS (diluted)",
]

BALANCE_SHEET_ROWS = [
    "Cash & Equivalents", "Accounts Receivable", "Inventory",
    "Other Current Assets", "Total Current Assets",
    "PP&E (net)", "Intangibles & Goodwill", "Other Long-term Assets",
    "Total Assets", "Accounts Payable", "Short-term Debt",
    "Other Current Liabilities", "Total Current Liabilities",
    "Long-term Debt", "Other Long-term Liabilities", "Total Liabilities",
    "Common Equity", "Retained Earnings", "Total Shareholders Equity",
    "Total Liabilities + Equity",
]

CASH_FLOW_ROWS = [
    "Net Income", "D&A", "Stock-based Compensation",
    "Change in Accounts Receivable", "Change in Inventory",
    "Change in Accounts Payable", "Other Working Capital Changes",
    "Operating Cash Flow", "Capex", "Free Cash Flow", "FCF Margin %",
    "Acquisitions", "Other Investing Activities", "Net Cash from Investing",
    "Debt Issuance / (Repayment)", "Share Repurchases", "Dividends",
    "Net Cash from Financing", "Net Change in Cash",
    "Beginning Cash", "Ending Cash",
]

SCENARIOS_METRICS = [
    "Revenue", "Revenue Growth %", "EBITDA", "EBITDA Margin %",
    "EPS (diluted)", "Free Cash Flow", "FCF Margin %",
    "Implied Price Target",
]

# Map from financial_metrics metric_name → canonical row name (case-insensitive)
_METRIC_MAP: dict[str, str] = {
    "total_revenue": "Revenue",
    "cost_of_revenue": "Cost of Revenue",
    "gross_profit": "Gross Profit",
    "operating_income": "EBIT",
    "ebitda": "EBITDA",
    "net_income": "Net Income",
    "diluted_eps": "EPS (diluted)",
    "basic_eps": "EPS (basic)",
    "operating_expense": "Total Operating Expenses",
    "research_and_development": "R&D",
    "selling_general_and_administrative": "G&A",
    "interest_expense": "Interest Expense",
    "income_tax_expense": "Income Tax",
    "total_assets": "Total Assets",
    "total_liabilities_net_minority_interest": "Total Liabilities",
    "stockholders_equity": "Total Shareholders Equity",
    "total_debt": "Long-term Debt",
    "cash_and_cash_equivalents": "Cash & Equivalents",
    "operating_cash_flow": "Operating Cash Flow",
    "free_cash_flow": "Free Cash Flow",
    "capital_expenditure": "Capex",
    "investing_cash_flow": "Net Cash from Investing",
    "financing_cash_flow": "Net Cash from Financing",
    "accounts_receivable": "Accounts Receivable",
    "inventory": "Inventory",
    "accounts_payable": "Accounts Payable",
}


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


class FinancialModelBuilder:
    """Builds a canonical 7-sheet financial model Excel file."""

    def __init__(
        self,
        ticker: str,
        historical_data: list[dict],
        estimates: list[dict],
        peers_data: dict[str, list[dict]],
        prior_assumptions: list[dict] | None,
        key_kpis: list[str],
        base_assumptions: dict | None,
        current_price: float | None,
    ):
        self.ticker = ticker
        self.historical_data = historical_data
        self.estimates = estimates
        self.peers_data = peers_data
        self.prior_assumptions = prior_assumptions or []
        self.key_kpis = key_kpis
        self.base_assumptions = base_assumptions or {}
        self.current_price = current_price

        self.sheets_generated: list[str] = []
        self._unmapped_metrics: list[str] = []
        self._assumptions_dict: list[dict] = []

        # Compute periods
        now = datetime.now()
        cur_year = now.year
        self.hist_annual = [f"FY{cur_year - 3}", f"FY{cur_year - 2}", f"FY{cur_year - 1}"]
        self.hist_quarterly = [f"Q{q}_{cur_year - 1}" for q in range(1, 5)]
        self.fwd_annual = [f"FY{cur_year}E", f"FY{cur_year + 1}E", f"FY{cur_year + 2}E"]
        self.fwd_quarterly = [f"Q{q}_{cur_year}E" for q in range(1, 5)]
        self.scenarios = ["Base", "Bull", "Bear"]

        # Index historical data for fast lookup
        self._hist_index: dict[tuple[str, str], float] = {}
        mapped_metrics: set[str] = set()
        for row in historical_data:
            metric_name = row.get("metric_name", "")
            canonical = _METRIC_MAP.get(metric_name)
            if canonical:
                period = str(row.get("period_end", ""))
                val = _safe_float(row.get("value"))
                if val is not None:
                    self._hist_index[(canonical, period)] = val
                    mapped_metrics.add(metric_name)

        # Track unmapped
        all_metric_names = {row.get("metric_name", "") for row in historical_data}
        self._unmapped_metrics = sorted(all_metric_names - mapped_metrics - {""})

    def build(self) -> tuple[bytes, list[str]]:
        """Build the Excel file. Returns (file_bytes, unmapped_metrics)."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        self._build_assumptions_sheet(wb)
        self._build_income_statement_sheet(wb)
        self._build_balance_sheet_sheet(wb)
        self._build_cash_flow_sheet(wb)
        self._build_valuation_sheet(wb)
        self._build_kpis_sheet(wb)
        self._build_scenarios_sheet(wb)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), self._unmapped_metrics

    def get_assumptions_dict(self) -> list[dict]:
        """Return assumptions as list of dicts for storage in assumptions_snapshot."""
        return self._assumptions_dict

    # ==================================================================
    # Sheet builders
    # ==================================================================

    def _build_assumptions_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("ASSUMPTIONS")
        self.sheets_generated.append("ASSUMPTIONS")

        # Column layout: Metric | hist_annual... | hist_quarterly... | fwd_annual (Base|Bull|Bear) | fwd_quarterly (Base only)
        headers = ["Metric"] + self.hist_annual + self.hist_quarterly
        for period in self.fwd_annual:
            for scenario in self.scenarios:
                headers.append(f"{period} {scenario}")
        for period in self.fwd_quarterly:
            headers.append(f"{period} Base")

        self._write_header_row(ws, headers)

        for row_idx, metric in enumerate(ASSUMPTIONS_ROWS, start=2):
            ws.cell(row=row_idx, column=1, value=metric).font = _BOLD

            # Compute default values from historical data
            default_val = self._compute_assumption_default(metric)

            # Check prior assumptions
            prior_val = self._get_prior_assumption(metric, "base")

            # Check user overrides
            override_val = self.base_assumptions.get(metric)

            base_val = override_val if override_val is not None else (prior_val if prior_val is not None else default_val)
            bull_val = base_val * 1.015 if base_val else None  # +150bps
            bear_val = base_val * 0.985 if base_val else None  # -150bps

            # Store for snapshot
            self._assumptions_dict.append({
                "metric": metric,
                "base": base_val,
                "bull": bull_val,
                "bear": bear_val,
            })

            col = 2
            # Historical annual (blank for assumptions — assumptions are forward-looking)
            for _ in self.hist_annual:
                col += 1
            # Historical quarterly
            for _ in self.hist_quarterly:
                col += 1

            # Forward annual: Base, Bull, Bear
            for _ in self.fwd_annual:
                for i, val in enumerate([base_val, bull_val, bear_val]):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.number_format = _FMT_PCT if "%" in metric or "Rate" in metric else _FMT_DOLLAR
                    cell.fill = [_BASE_FILL, _BULL_FILL, _BEAR_FILL][i]
                    col += 1

            # Forward quarterly: Base only
            for _ in self.fwd_quarterly:
                cell = ws.cell(row=row_idx, column=col, value=base_val)
                cell.number_format = _FMT_PCT if "%" in metric or "Rate" in metric else _FMT_DOLLAR
                cell.fill = _BASE_FILL
                col += 1

        self._format_sheet(ws, len(headers))

    def _build_income_statement_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("INCOME_STATEMENT")
        self.sheets_generated.append("INCOME_STATEMENT")
        self._build_financial_sheet(ws, INCOME_STATEMENT_ROWS)

    def _build_balance_sheet_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("BALANCE_SHEET")
        self.sheets_generated.append("BALANCE_SHEET")
        self._build_financial_sheet(ws, BALANCE_SHEET_ROWS)

    def _build_cash_flow_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("CASH_FLOW")
        self.sheets_generated.append("CASH_FLOW")
        self._build_financial_sheet(ws, CASH_FLOW_ROWS)

    def _build_financial_sheet(self, ws, rows: list[str]) -> None:
        """Build IS, BS, or CF sheet with historical + forward columns."""
        headers = ["Metric"] + self.hist_annual + self.hist_quarterly
        for period in self.fwd_annual:
            for scenario in self.scenarios:
                headers.append(f"{period} {scenario}")
        for period in self.fwd_quarterly:
            headers.append(f"{period} Base")

        self._write_header_row(ws, headers)

        for row_idx, metric in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=metric).font = _BOLD

            is_pct = "%" in metric or "Margin" in metric
            fmt = _FMT_PCT if is_pct else (_FMT_PER_SHARE if "EPS" in metric else _FMT_DOLLAR)

            col = 2
            # Historical annual
            for period in self.hist_annual:
                val = self._lookup_historical(metric, period)
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = fmt
                col += 1

            # Historical quarterly
            for period in self.hist_quarterly:
                val = self._lookup_historical(metric, period)
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = fmt
                col += 1

            # Forward annual: Base, Bull, Bear
            for _ in self.fwd_annual:
                for i in range(3):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = [_BASE_FILL, _BULL_FILL, _BEAR_FILL][i]
                    cell.number_format = fmt
                    col += 1

            # Forward quarterly: Base only
            for _ in self.fwd_quarterly:
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = _BASE_FILL
                cell.number_format = fmt
                col += 1

        self._format_sheet(ws, len(headers))

    def _build_valuation_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("VALUATION")
        self.sheets_generated.append("VALUATION")

        # --- DCF Section ---
        ws.cell(row=1, column=1, value="DCF Valuation (Base Case)").font = Font(bold=True, size=12)

        dcf_rows = [
            "Projected FCF Year 1", "Projected FCF Year 2", "Projected FCF Year 3",
            "Projected FCF Year 4", "Projected FCF Year 5",
            "Terminal Value", "Enterprise Value", "Net Debt",
            "Equity Value", "Diluted Shares (M)", "Implied Share Price",
            "Current Price", "Upside / Downside %",
        ]

        wacc = self._get_assumption_val("WACC (%)", "base") or 0.10
        terminal_growth = self._get_assumption_val("Terminal Growth Rate (%)", "base") or 0.025

        # Compute simple DCF placeholder values
        last_fcf = self._lookup_historical("Free Cash Flow", self.hist_annual[-1]) or 0
        diluted_shares = self._get_assumption_val("Share Count - Diluted (M)", "base") or 1

        fcf_projections = []
        fcf = last_fcf
        for yr in range(1, 6):
            fcf = fcf * 1.05  # Simple 5% growth placeholder
            fcf_projections.append(fcf)

        terminal_val = (fcf_projections[-1] * (1 + terminal_growth) / (wacc - terminal_growth)) if wacc > terminal_growth else 0
        pv_sum = sum(f / (1 + wacc) ** i for i, f in enumerate(fcf_projections, 1))
        pv_terminal = terminal_val / (1 + wacc) ** 5
        ev = pv_sum + pv_terminal

        net_debt = (self._lookup_historical("Long-term Debt", self.hist_annual[-1]) or 0) - \
                   (self._lookup_historical("Cash & Equivalents", self.hist_annual[-1]) or 0)
        equity_val = ev - net_debt
        implied_price = equity_val / diluted_shares if diluted_shares else 0
        upside = ((implied_price - (self.current_price or 0)) / (self.current_price or 1)) * 100 if self.current_price else 0

        dcf_values = fcf_projections + [
            terminal_val, ev, net_debt, equity_val,
            diluted_shares, implied_price, self.current_price, upside,
        ]

        for row_idx, (label, val) in enumerate(zip(dcf_rows, dcf_values), start=3):
            ws.cell(row=row_idx, column=1, value=label).font = _BOLD
            cell = ws.cell(row=row_idx, column=2, value=val)
            cell.number_format = _FMT_PCT if "%" in label else _FMT_DOLLAR

        # --- Comps Section ---
        comps_start = len(dcf_rows) + 5
        ws.cell(row=comps_start, column=1, value="Peer Comparables").font = Font(bold=True, size=12)

        comp_headers = ["Ticker", "EV/EBITDA NTM", "P/E NTM", "EV/Sales NTM", "FCF Yield"]
        for col_idx, h in enumerate(comp_headers, 1):
            cell = ws.cell(row=comps_start + 1, column=col_idx, value=h)
            cell.font = _BOLD
            cell.fill = _HEADER_FILL

        row_idx = comps_start + 2
        for peer_ticker, peer_data in self.peers_data.items():
            ws.cell(row=row_idx, column=1, value=peer_ticker)
            # Compute simple multiples from peer data if available
            for col_idx in range(2, 6):
                ws.cell(row=row_idx, column=col_idx, value="N/A")
            row_idx += 1

        # Subject company row
        ws.cell(row=row_idx, column=1, value=f"{self.ticker} (Subject)").font = _BOLD
        for col_idx in range(2, 6):
            ws.cell(row=row_idx, column=col_idx, value="N/A")

        # Formatting
        ws.column_dimensions["A"].width = 30
        for col in range(2, 6):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "B1"

    def _build_kpis_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("KPIs")
        self.sheets_generated.append("KPIs")

        headers = ["KPI"] + self.hist_annual + self.hist_quarterly + self.fwd_annual + self.fwd_quarterly
        self._write_header_row(ws, headers)

        kpis = self.key_kpis if self.key_kpis else ["Add company-specific KPIs here"]

        for row_idx, kpi in enumerate(kpis, start=2):
            ws.cell(row=row_idx, column=1, value=kpi).font = _BOLD
            # Try to fill historical values if metric exists
            col = 2
            for period in self.hist_annual + self.hist_quarterly:
                val = self._lookup_historical(kpi, period)
                if val is not None:
                    ws.cell(row=row_idx, column=col, value=val).number_format = _FMT_DOLLAR
                col += 1
            # Forward periods: empty for analyst to fill
            for _ in self.fwd_annual + self.fwd_quarterly:
                col += 1

        self._format_sheet(ws, len(headers))

    def _build_scenarios_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("SCENARIOS")
        self.sheets_generated.append("SCENARIOS")

        # Three blocks side by side
        all_metrics = SCENARIOS_METRICS + [f"KPI: {k}" for k in self.key_kpis]

        for block_idx, scenario in enumerate(self.scenarios):
            col_start = 1 + block_idx * (len(self.fwd_annual) + 1)

            # Scenario header
            cell = ws.cell(row=1, column=col_start, value=f"{scenario} Case")
            cell.font = Font(bold=True, size=12)
            fill = [_BASE_FILL, _BULL_FILL, _BEAR_FILL][block_idx]
            cell.fill = fill

            # Period headers
            ws.cell(row=2, column=col_start, value="Metric").font = _BOLD
            for j, period in enumerate(self.fwd_annual, 1):
                cell = ws.cell(row=2, column=col_start + j, value=period)
                cell.font = _BOLD
                cell.fill = fill

            # Metric rows
            for row_idx, metric in enumerate(all_metrics, start=3):
                ws.cell(row=row_idx, column=col_start, value=metric).font = _BOLD
                is_pct = "%" in metric or "Margin" in metric
                fmt = _FMT_PCT if is_pct else (_FMT_PER_SHARE if "EPS" in metric else _FMT_DOLLAR)
                for j in range(len(self.fwd_annual)):
                    cell = ws.cell(row=row_idx, column=col_start + j + 1)
                    cell.number_format = fmt
                    cell.fill = fill

        # Formatting
        total_cols = 3 * (len(self.fwd_annual) + 1)
        ws.column_dimensions["A"].width = 30
        for col in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "B1"

    # ==================================================================
    # Helpers
    # ==================================================================

    def _write_header_row(self, ws, headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _BOLD
            cell.fill = _HEADER_FILL

    def _format_sheet(self, ws, num_cols: int) -> None:
        ws.column_dimensions["A"].width = 30
        for col in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "B2"

    def _lookup_historical(self, metric: str, period: str) -> float | None:
        """Look up a historical value from the indexed data."""
        val = self._hist_index.get((metric, period))
        if val is not None:
            return val
        # Try fuzzy match on period (FY2024 vs 2024-12-31)
        for (m, p), v in self._hist_index.items():
            if m == metric and period in p:
                return v
        return None

    def _compute_assumption_default(self, metric: str) -> float | None:
        """Compute a default assumption value from last 3 years of historical data."""
        # Extract relevant historical values
        if "Revenue Growth" in metric:
            revs = []
            for period in self.hist_annual:
                val = self._lookup_historical("Revenue", period)
                if val:
                    revs.append(val)
            if len(revs) >= 2:
                return (revs[-1] / revs[-2]) - 1.0
            return 0.05  # 5% default
        elif "Gross Margin" in metric:
            return self._median_ratio("Gross Profit", "Revenue")
        elif "EBITDA Margin" in metric:
            return self._median_ratio("EBITDA", "Revenue")
        elif "R&D" in metric:
            return self._median_ratio("R&D", "Revenue")
        elif "S&M" in metric:
            return 0.10  # 10% default
        elif "G&A" in metric:
            return 0.05  # 5% default
        elif "Tax Rate" in metric:
            return 0.21  # 21% default
        elif "D&A" in metric:
            return 0.03  # 3% default
        elif "Capex" in metric:
            return 0.04  # 4% default
        elif "Days Sales" in metric:
            return 45.0
        elif "Days Inventory" in metric:
            return 30.0
        elif "Days Payable" in metric:
            return 40.0
        elif "Share Count" in metric and "Basic" in metric:
            return self._lookup_historical("Shares Outstanding (basic)", self.hist_annual[-1]) or 1000
        elif "Share Count" in metric and "Diluted" in metric:
            return self._lookup_historical("Shares Outstanding (diluted)", self.hist_annual[-1]) or 1000
        elif "WACC" in metric:
            return 0.10
        elif "Terminal Growth" in metric:
            return 0.025
        return None

    def _median_ratio(self, numerator: str, denominator: str) -> float | None:
        """Compute median ratio of numerator/denominator over historical annual periods."""
        ratios = []
        for period in self.hist_annual:
            num = self._lookup_historical(numerator, period)
            den = self._lookup_historical(denominator, period)
            if num is not None and den and den != 0:
                ratios.append(num / den)
        if ratios:
            ratios.sort()
            mid = len(ratios) // 2
            return ratios[mid]
        return None

    def _get_prior_assumption(self, metric: str, scenario: str) -> float | None:
        """Look up a value from prior model assumptions."""
        for row in self.prior_assumptions:
            if row.get("metric") == metric and row.get("scenario") == scenario:
                return _safe_float(row.get("value"))
        return None

    def _get_assumption_val(self, metric: str, scenario: str) -> float | None:
        """Get assumption value from the computed assumptions dict."""
        for row in self._assumptions_dict:
            if row.get("metric") == metric:
                return _safe_float(row.get(scenario))
        return None
