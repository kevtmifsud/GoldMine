#!/usr/bin/env python3
"""Model processing job: reads Excel models from S3, writes to model_outputs and internal_estimates.

S3 path convention: models/{ticker}/{ticker}_model_v{N}_{date}.xlsx
Processes the highest version number per ticker.

Usage:
    python scripts/model_processing_job.py                     # All tickers with models in S3
    python scripts/model_processing_job.py --tickers AAPL MSFT # Specific tickers
    python scripts/model_processing_job.py --batch-size 20     # Custom batch size
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import tempfile
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl
import psycopg2
import structlog
from dotenv import load_dotenv
from psycopg2.extras import execute_values

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / "backend" / ".env")

DATABASE_URL = os.environ.get("SUPABASE_DATABASE_URL", "")
S3_BUCKET = os.environ.get("GOLDMINE_S3_BUCKET", "")
AWS_REGION = os.environ.get("GOLDMINE_AWS_REGION", "us-east-1")

log = structlog.get_logger()

# ---------------------------------------------------------------------------
# Canonical sheet names → db sheet values
# ---------------------------------------------------------------------------

SHEET_DB_NAMES = {
    "ASSUMPTIONS": "assumptions",
    "INCOME_STATEMENT": "income_statement",
    "BALANCE_SHEET": "balance_sheet",
    "CASH_FLOW": "cash_flow",
    "VALUATION": "valuation",
    "KPIs": "kpis",
    "SCENARIOS": "scenarios",
}

# Canonical metric names per sheet (from models.md).  Metric names in the
# Excel must match these exactly (case-insensitive comparison).  Anything not
# in this set is logged as unmapped and skipped.

ASSUMPTIONS_METRICS = {
    "revenue_growth_rate", "revenue_by_segment", "gross_margin_pct",
    "rd_pct_revenue", "sm_pct_revenue", "ga_pct_revenue",
    "ebitda_margin_pct", "effective_tax_rate", "da_pct_revenue",
    "capex_pct_revenue", "days_sales_outstanding", "days_inventory_outstanding",
    "days_payable_outstanding", "share_count_basic", "share_count_diluted",
    "wacc", "terminal_growth_rate",
}

INCOME_STATEMENT_METRICS = {
    "revenue", "revenue_by_segment", "cost_of_revenue", "gross_profit",
    "gross_margin_pct", "rd", "sm", "ga", "total_operating_expenses",
    "ebit", "ebit_margin_pct", "da", "ebitda", "ebitda_margin_pct",
    "interest_expense", "other_income_expense", "pre_tax_income",
    "income_tax", "net_income", "net_margin_pct",
    "shares_outstanding_basic", "shares_outstanding_diluted",
    "eps_basic", "eps_diluted",
}

BALANCE_SHEET_METRICS = {
    "cash_and_equivalents", "accounts_receivable", "inventory",
    "other_current_assets", "total_current_assets", "ppe_net",
    "intangibles_and_goodwill", "other_long_term_assets", "total_assets",
    "accounts_payable", "short_term_debt", "other_current_liabilities",
    "total_current_liabilities", "long_term_debt", "other_long_term_liabilities",
    "total_liabilities", "common_equity", "retained_earnings",
    "total_shareholders_equity", "total_liabilities_and_equity",
}

CASH_FLOW_METRICS = {
    "net_income", "da", "stock_based_compensation",
    "change_in_accounts_receivable", "change_in_inventory",
    "change_in_accounts_payable", "other_working_capital_changes",
    "operating_cash_flow", "capex", "free_cash_flow", "fcf_margin_pct",
    "acquisitions", "other_investing_activities", "net_cash_from_investing",
    "debt_issuance_repayment", "share_repurchases", "dividends",
    "net_cash_from_financing", "net_change_in_cash",
    "beginning_cash", "ending_cash",
}

VALUATION_METRICS = {
    "projection_period_fcf", "terminal_value", "enterprise_value",
    "net_debt", "equity_value", "implied_share_price",
    "current_price", "upside_downside_pct",
    # Comps metrics
    "ev_ebitda_ntm", "pe_ntm", "ev_sales_ntm", "fcf_yield",
    "implied_multiples", "premium_discount_to_peer_median",
    # Price target summary
    "price_target",
}

KPIS_METRICS: set[str] = set()  # KPIs are ticker-specific — accept all

SCENARIOS_METRICS = {
    "revenue", "revenue_growth_pct", "ebitda", "ebitda_margin_pct",
    "eps", "fcf", "implied_price_target",
    # Plus ticker-specific KPIs — we accept anything on the SCENARIOS sheet
}

SHEET_METRICS = {
    "ASSUMPTIONS": ASSUMPTIONS_METRICS,
    "INCOME_STATEMENT": INCOME_STATEMENT_METRICS,
    "BALANCE_SHEET": BALANCE_SHEET_METRICS,
    "CASH_FLOW": CASH_FLOW_METRICS,
    "VALUATION": VALUATION_METRICS,
    "KPIs": KPIS_METRICS,
    "SCENARIOS": SCENARIOS_METRICS,
}

# Metrics that should produce internal_estimates from the SCENARIOS sheet
ESTIMATE_METRICS = {
    "revenue", "revenue_growth_pct", "ebitda", "ebitda_margin_pct",
    "eps", "fcf",
}

# Units we infer from metric names
UNIT_MAP = {
    "pct": "percentage",
    "margin": "percentage",
    "growth": "percentage",
    "rate": "percentage",
    "eps": "per_share",
    "per_share": "per_share",
    "ratio": "ratio",
    "yield": "percentage",
}


def _infer_unit(metric: str) -> str:
    """Infer the unit for a metric based on its name."""
    lower = metric.lower()
    for keyword, unit in UNIT_MAP.items():
        if keyword in lower:
            return unit
    return "dollars"


# ---------------------------------------------------------------------------
# S3 helpers
# ---------------------------------------------------------------------------

def get_s3_client():
    """Create S3 client via Supabase storage utility — never instantiate boto3 directly."""
    sys.path.insert(0, str(ROOT / "backend"))
    from app.workflows.storage import get_s3_client as _get_client
    return _get_client()


def list_model_tickers(s3_client) -> list[str]:
    """List all tickers that have models in S3."""
    tickers = set()
    paginator = s3_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=S3_BUCKET, Prefix="models/", Delimiter="/"):
        for prefix in page.get("CommonPrefixes", []):
            # prefix looks like "models/AAPL/"
            parts = prefix["Prefix"].rstrip("/").split("/")
            if len(parts) == 2:
                tickers.add(parts[1])
    return sorted(tickers)


def find_latest_model(s3_client, ticker: str) -> tuple[str, str] | None:
    """Find the S3 key of the highest-version model for a ticker.

    Returns (s3_key, version_str) or None if no models found.
    """
    prefix = f"models/{ticker}/"
    pattern = re.compile(
        rf"^models/{re.escape(ticker)}/{re.escape(ticker)}_model_v(\d+)_(\d{{4}}-\d{{2}}-\d{{2}})\.xlsx$"
    )

    best_version = -1
    best_key = None

    paginator = s3_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=prefix):
        for obj in page.get("Contents", []):
            m = pattern.match(obj["Key"])
            if m:
                version_num = int(m.group(1))
                if version_num > best_version:
                    best_version = version_num
                    best_key = obj["Key"]

    if best_key is None:
        return None
    return best_key, str(best_version)


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _normalize_metric(raw: str) -> str:
    """Convert a human-readable metric label to snake_case canonical name."""
    s = raw.strip()
    s = re.sub(r"[%()]", "", s)
    s = re.sub(r"[/&,\-]+", "_", s)
    s = re.sub(r"\s+", "_", s)
    s = s.lower().strip("_")
    # Collapse multiple underscores
    s = re.sub(r"_+", "_", s)
    return s


def _is_forward_period(period: str) -> bool:
    """Check if a period string represents a forward/forecast period."""
    today = date.today()
    # Extract year from period like 2026Q1 or 2026A
    m = re.match(r"(\d{4})", period)
    if m:
        year = int(m.group(1))
        return year >= today.year
    return False


def parse_sheet(ws, sheet_name: str, allowed_metrics: set[str]) -> list[dict]:
    """Parse a worksheet into (metric, period, scenario, value, unit) tuples.

    Expected layout:
      Row 1: header row with period labels (e.g. 2022A, 2023Q1, etc.)
      Row 2: scenario labels (base, bull, bear) — optional, defaults to 'actual' for historical
      Row 3+: metric name in column A, values in subsequent columns

    Returns list of dicts with keys: metric, period, scenario, value, unit.
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return []

    # Find the header row (first row with period-looking values)
    header_row = rows[0]
    scenario_row = rows[1] if len(rows) > 1 else None

    # Build column map: col_index -> (period, scenario)
    col_map: dict[int, tuple[str, str]] = {}
    for col_idx, val in enumerate(header_row):
        if col_idx == 0 or val is None:
            continue
        period = str(val).strip()
        # Determine scenario from scenario row, default to 'actual' for historical
        scenario = "actual"
        if scenario_row and col_idx < len(scenario_row) and scenario_row[col_idx]:
            scenario_val = str(scenario_row[col_idx]).strip().lower()
            if scenario_val in ("base", "bull", "bear"):
                scenario = scenario_val
            elif _is_forward_period(period) and scenario_val not in ("actual",):
                scenario = "base"  # Default forward to base if no scenario label
        elif _is_forward_period(period):
            scenario = "base"
        col_map[col_idx] = (period, scenario)

    # Determine data start row (skip header + optional scenario row)
    data_start = 2 if (scenario_row and any(
        str(v).strip().lower() in ("base", "bull", "bear", "actual", "scenario")
        for v in scenario_row if v is not None
    )) else 1

    results = []
    for row in rows[data_start:]:
        if not row or not row[0]:
            continue
        raw_metric = str(row[0]).strip()
        metric = _normalize_metric(raw_metric)

        if not metric:
            continue

        # Check if metric is canonical (KPIs and SCENARIOS accept all)
        if allowed_metrics and metric not in allowed_metrics:
            log.debug("unmapped_metric", sheet=sheet_name, raw=raw_metric, normalized=metric)
            continue

        unit = _infer_unit(metric)

        for col_idx, (period, scenario) in col_map.items():
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx]
            if cell_val is None:
                continue
            try:
                value = float(cell_val)
            except (ValueError, TypeError):
                continue

            results.append({
                "metric": metric,
                "period": period,
                "scenario": scenario,
                "value": value,
                "unit": unit,
            })

    return results


def parse_workbook(file_path: str) -> dict[str, list[dict]]:
    """Parse all canonical sheets from a workbook.

    Returns dict of sheet_name -> list of parsed metric rows.
    Missing sheets are logged and skipped.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    results = {}

    for excel_name, db_name in SHEET_DB_NAMES.items():
        if excel_name not in wb.sheetnames:
            log.warning("missing_sheet", sheet=excel_name)
            continue
        ws = wb[excel_name]
        allowed = SHEET_METRICS.get(excel_name, set())
        parsed = parse_sheet(ws, excel_name, allowed)
        results[db_name] = parsed
        log.info("parsed_sheet", sheet=excel_name, rows=len(parsed))

    wb.close()
    return results


# ---------------------------------------------------------------------------
# Database writes
# ---------------------------------------------------------------------------

def get_conn():
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = True
    return conn


def write_model_outputs(
    conn,
    ticker: str,
    version: str,
    as_of: date,
    sheet_data: dict[str, list[dict]],
) -> int:
    """Write parsed sheet data to model_outputs. Returns total rows inserted."""
    total = 0
    with conn.cursor() as cur:
        for sheet_name, rows in sheet_data.items():
            if not rows:
                continue
            values = [
                (
                    str(uuid.uuid4()),
                    ticker,
                    version,
                    as_of,
                    sheet_name,
                    r["metric"],
                    r["period"],
                    r["scenario"],
                    r["value"],
                    r["unit"],
                    "system",
                    datetime.utcnow(),
                )
                for r in rows
            ]
            execute_values(
                cur,
                """INSERT INTO model_outputs
                   (id, ticker, version, as_of_date, sheet, metric, period,
                    scenario, value, unit, created_by, created_at)
                   VALUES %s""",
                values,
                page_size=500,
            )
            total += len(values)
    return total


def write_internal_estimates_from_scenarios(
    conn,
    ticker: str,
    version: str,
    scenario_rows: list[dict],
) -> int:
    """Extract forward estimates from SCENARIOS sheet and write to internal_estimates."""
    values = []
    for row in scenario_rows:
        if not _is_forward_period(row["period"]):
            continue
        # Only write estimate metrics (revenue, eps, ebitda, fcf, etc.)
        if row["metric"] not in ESTIMATE_METRICS:
            # Also accept KPI metrics (anything not in the skip list)
            if row["metric"] in ("implied_price_target",):
                continue  # Price targets handled separately from VALUATION
            # Accept ticker-specific KPIs
            if row["metric"] in SCENARIOS_METRICS and row["metric"] not in ESTIMATE_METRICS:
                continue
            # Otherwise treat as a KPI estimate
        values.append((
            str(uuid.uuid4()),
            ticker,
            row["period"],
            row["metric"],
            row["value"],
            row["unit"],
            version,
            datetime.utcnow(),
        ))

    if not values:
        return 0

    with conn.cursor() as cur:
        execute_values(
            cur,
            """INSERT INTO internal_estimates
               (id, ticker, period, metric, value, unit,
                model_version, created_at)
               VALUES %s""",
            values,
            page_size=500,
        )
    return len(values)


def write_internal_estimates_price_targets(
    conn,
    ticker: str,
    version: str,
    valuation_rows: list[dict],
) -> int:
    """Extract price targets from VALUATION sheet and write to internal_estimates."""
    values = []
    for row in valuation_rows:
        if row["metric"] not in ("implied_share_price", "price_target"):
            continue
        if not _is_forward_period(row["period"]):
            continue
        values.append((
            str(uuid.uuid4()),
            ticker,
            row["period"],
            "price_target",
            row["value"],
            "per_share",
            version,
            datetime.utcnow(),
        ))

    if not values:
        return 0

    with conn.cursor() as cur:
        execute_values(
            cur,
            """INSERT INTO internal_estimates
               (id, ticker, period, metric, value, unit,
                model_version, created_at)
               VALUES %s""",
            values,
            page_size=500,
        )
    return len(values)


# ---------------------------------------------------------------------------
# Per-ticker processing
# ---------------------------------------------------------------------------

def process_ticker(s3_client, conn, ticker: str) -> dict:
    """Process a single ticker: download model from S3, parse, write to DB."""
    result = {"ticker": ticker, "status": "skipped", "model_output_rows": 0, "estimate_rows": 0}

    model_info = find_latest_model(s3_client, ticker)
    if model_info is None:
        log.info("no_model_found", ticker=ticker)
        return result

    s3_key, version = model_info
    log.info("processing_model", ticker=ticker, s3_key=s3_key, version=version)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
        s3_client.download_file(S3_BUCKET, s3_key, tmp.name)

        sheet_data = parse_workbook(tmp.name)

    if not sheet_data:
        log.warning("empty_workbook", ticker=ticker)
        result["status"] = "empty"
        return result

    as_of = date.today()

    # Write model_outputs
    output_rows = write_model_outputs(conn, ticker, version, as_of, sheet_data)
    result["model_output_rows"] = output_rows

    # Write internal_estimates from SCENARIOS
    estimate_rows = 0
    if "scenarios" in sheet_data:
        estimate_rows += write_internal_estimates_from_scenarios(
            conn, ticker, version, sheet_data["scenarios"]
        )

    # Write price targets from VALUATION
    if "valuation" in sheet_data:
        estimate_rows += write_internal_estimates_price_targets(
            conn, ticker, version, sheet_data["valuation"]
        )

    result["estimate_rows"] = estimate_rows
    result["status"] = "ok"
    log.info(
        "ticker_complete",
        ticker=ticker,
        model_output_rows=output_rows,
        estimate_rows=estimate_rows,
    )
    return result


# ---------------------------------------------------------------------------
# Async importable functions (called by workflow)
# ---------------------------------------------------------------------------

async def process_ticker_from_s3(
    ticker: str,
    version: str,
    s3_key: str,
) -> dict:
    """Process one ticker's model from S3.

    Downloads file, parses all 7 sheets, writes to model_outputs and
    internal_estimates. Returns summary dict with rows_written count.

    This is the async entry point used by FinancialModelGenerationWorkflow.
    """
    import tempfile

    sys.path.insert(0, str(ROOT / "backend"))
    from app.workflows.storage import download_model

    log.info("async_processing_model", ticker=ticker, version=version, s3_key=s3_key)

    file_bytes = download_model(s3_key)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
        tmp.write(file_bytes)
        tmp.flush()
        sheet_data = parse_workbook(tmp.name)

    if not sheet_data:
        return {"ticker": ticker, "version": version, "rows_written": 0, "estimates_written": 0}

    conn = get_conn()
    as_of = date.today()

    output_rows = write_model_outputs(conn, ticker, version, as_of, sheet_data)

    estimate_rows = 0
    if "scenarios" in sheet_data:
        estimate_rows += write_internal_estimates_from_scenarios(
            conn, ticker, version, sheet_data["scenarios"]
        )
    if "valuation" in sheet_data:
        estimate_rows += write_internal_estimates_price_targets(
            conn, ticker, version, sheet_data["valuation"]
        )

    conn.close()

    log.info("async_model_processed", ticker=ticker, version=version,
             rows_written=output_rows, estimates_written=estimate_rows)

    return {
        "ticker": ticker,
        "version": version,
        "rows_written": output_rows,
        "estimates_written": estimate_rows,
        "sheets_parsed": list(sheet_data.keys()),
    }


async def process_all_tickers() -> dict:
    """Process all tickers that have models in S3.

    Called by the nightly cron runner.
    Returns summary dict.
    """
    return run()


# ---------------------------------------------------------------------------
# Main (sync)
# ---------------------------------------------------------------------------

def run(tickers: list[str] | None = None, batch_size: int = 10) -> dict:
    """Run the model processing job.

    Args:
        tickers: Specific tickers to process. None = all tickers with models in S3.
        batch_size: Number of tickers to process per batch.

    Returns:
        Summary dict with processed/failed/skipped counts.
    """
    if not S3_BUCKET:
        raise RuntimeError("GOLDMINE_S3_BUCKET environment variable is required")
    if not DATABASE_URL:
        raise RuntimeError("SUPABASE_DATABASE_URL environment variable is required")

    s3_client = get_s3_client()

    if tickers is None:
        tickers = list_model_tickers(s3_client)
        log.info("discovered_tickers", count=len(tickers))

    if not tickers:
        log.info("no_tickers_to_process")
        return {"processed": 0, "failed": 0, "skipped": 0}

    conn = get_conn()
    processed = 0
    failed = 0
    skipped = 0

    # Process in batches
    for batch_start in range(0, len(tickers), batch_size):
        batch = tickers[batch_start : batch_start + batch_size]
        log.info(
            "processing_batch",
            batch_start=batch_start,
            batch_size=len(batch),
            total=len(tickers),
        )

        for ticker in batch:
            try:
                result = process_ticker(s3_client, conn, ticker)
                if result["status"] == "ok":
                    processed += 1
                else:
                    skipped += 1
            except Exception:
                log.exception("ticker_failed", ticker=ticker)
                failed += 1

    conn.close()

    summary = {"processed": processed, "failed": failed, "skipped": skipped}
    log.info("job_complete", **summary)
    return summary


def main():
    parser = argparse.ArgumentParser(description="Process Excel models from S3")
    parser.add_argument("--tickers", nargs="*", help="Specific tickers to process")
    parser.add_argument("--batch-size", type=int, default=10, help="Batch size (default 10)")
    args = parser.parse_args()

    summary = run(tickers=args.tickers, batch_size=args.batch_size)
    if summary["failed"] > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
